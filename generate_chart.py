import pandas as pd
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList

def criar_graficos_analise(caminho_arquivo_excel: str):
    """
    Lê uma planilha de respostas, analisa os dados e gera um novo arquivo Excel
    com os dados brutos e gráficos de análise.

    Args:
        caminho_arquivo_excel: O caminho para o arquivo .xlsx de entrada.
    """
    try:
        df = pd.read_excel(caminho_arquivo_excel)
        # Corrige um possível erro de digitação no nome da coluna
        if 'reposta correta' in df.columns:
            df.rename(columns={'reposta correta': 'is_correct'}, inplace=True)
            
    except FileNotFoundError:
        print(f"Erro: O arquivo '{caminho_arquivo_excel}' não foi encontrado.")
        return
    except Exception as e:
        print(f"Ocorreu um erro ao ler o arquivo: {e}")
        return

    arquivo_saida = 'analise_de_respostas_com_graficos.xlsx'
    
    with pd.ExcelWriter(arquivo_saida, engine='openpyxl') as writer:
        # --- Passo 1: Salvar os dados originais em uma aba ---
        df.to_excel(writer, sheet_name='Dados Brutos', index=False)
        workbook = writer.book

        # --- Passo 2: Análise e Gráfico de Acurácia Geral (Pizza) ---
        ws_geral = workbook.create_sheet(title='Resumo Acurácia Geral')
        acuracia_geral = df['is_correct'].value_counts().reset_index()
        acuracia_geral.columns = ['Resultado', 'Contagem']
        for r_idx, row in enumerate(acuracia_geral.iterrows(), 2):
            for c_idx, value in enumerate(row[1], 1):
                ws_geral.cell(row=r_idx, column=c_idx, value=value)
        ws_geral.cell(row=1, column=1, value='Resultado')
        ws_geral.cell(row=1, column=2, value='Contagem')
        
        pie = PieChart()
        labels = Reference(ws_geral, min_col=1, min_row=2, max_row=len(acuracia_geral) + 1)
        data = Reference(ws_geral, min_col=2, min_row=1, max_row=len(acuracia_geral) + 1)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "Acurácia Geral (Acertos vs. Erros)"
        pie.data_labels = DataLabelList(showPercent=True)
        ws_geral.add_chart(pie, "E2")

        # --- Passo 3: Análise e Gráfico por Tipo de Pergunta (Barras) ---
        ws_tipo = workbook.create_sheet(title='Análise por Tipo')
        acuracia_por_tipo = df.groupby('type')['is_correct'].value_counts().unstack(fill_value=0)
        acuracia_por_tipo.reset_index(inplace=True)
        
        # Escrevendo os dados na planilha para usar como fonte do gráfico
        ws_tipo.append(acuracia_por_tipo.columns.tolist())
        for r_idx, row in enumerate(acuracia_por_tipo.iterrows(), 2):
            for c_idx, value in enumerate(row[1], 1):
                ws_tipo.cell(row=r_idx, column=c_idx, value=value)

        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Desempenho por Tipo de Pergunta"
        chart.y_axis.title = 'Número de Respostas'
        chart.x_axis.title = 'Tipo de Pergunta'
        
        data = Reference(ws_tipo, min_col=2, min_row=1, max_row=len(acuracia_por_tipo) + 1, max_col=len(acuracia_por_tipo.columns))
        cats = Reference(ws_tipo, min_col=1, min_row=2, max_row=len(acuracia_por_tipo) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws_tipo.add_chart(chart, "F2")

        # --- Passo 4: Análise e Gráfico por Duração do Vídeo (Barras) ---
        ws_duracao = workbook.create_sheet(title='Análise por Duração')
        acuracia_por_duracao = df.groupby('lenght')['is_correct'].value_counts(normalize=True).mul(100).unstack(fill_value=0)
        acuracia_por_duracao.reset_index(inplace=True)
        
        ws_duracao.append(acuracia_por_duracao.columns.tolist())
        for r_idx, row in enumerate(acuracia_por_duracao.iterrows(), 2):
            for c_idx, value in enumerate(row[1], 1):
                ws_duracao.cell(row=r_idx, column=c_idx, value=value)
        
        chart2 = BarChart()
        chart2.title = "Taxa de Acerto (%) por Duração do Segmento"
        chart2.y_axis.title = 'Taxa de Acerto (%)'
        chart2.x_axis.title = 'Duração'
        
        data = Reference(ws_duracao, min_col=2, min_row=1, max_row=len(acuracia_por_duracao) + 1, max_col=len(acuracia_por_duracao.columns))
        cats = Reference(ws_duracao, min_col=1, min_row=2, max_row=len(acuracia_por_duracao) + 1)
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
        ws_duracao.add_chart(chart2, "F2")

    print(f"Análise concluída. Arquivo '{arquivo_saida}' gerado com sucesso!")

# --- Como executar o script ---
if __name__ == "__main__":
    # Coloque o nome do seu arquivo Excel aqui
    nome_do_arquivo = 'responses/responses_gemini-2.0-flash.xlsx' # Substitua pelo nome real do seu arquivo
    criar_graficos_analise(nome_do_arquivo)